import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Du måste installera openpyxl först, t.ex.:", file=sys.stderr)
    print("  (.venv) pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Konvertera Ahlsell EL_Excel.xlsx till normaliserad price_catalog_ahlsell.json"
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Sökväg till EL_Excel.xlsx från Ahlsell",
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Sökväg till output JSON-fil (t.ex. price_catalog_ahlsell.json)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        print(f"Hittar inte input-fil: {input_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Läser Excel-fil: {input_path}")
    wb = load_workbook(input_path, read_only=True, data_only=True)

    # Försök använda bladet "Excel", annars första bladet
    if "Excel" in wb.sheetnames:
        ws = wb["Excel"]
    else:
        ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)

    try:
        header = next(rows_iter)
    except StopIteration:
        print("Excel-filen verkar vara tom", file=sys.stderr)
        sys.exit(1)

    # Bygg en map: kolumnnamn -> index
    col_map = {}
    for idx, name in enumerate(header):
        if isinstance(name, str):
            key = name.strip()
            if key:
                col_map[key] = idx

    required_cols = ["Artikelnr", "GNP", "Benämning", "Materialklass", "Enhet"]
    missing = [c for c in required_cols if c not in col_map]
    if missing:
        print("Saknar förväntade kolumner i Excel-filen:", ", ".join(missing), file=sys.stderr)
        print("Hittade kolumner:", ", ".join(col_map.keys()), file=sys.stderr)
        sys.exit(1)

    idx_art = col_map["Artikelnr"]
    idx_gnp = col_map["GNP"]
    idx_ben = col_map["Benämning"]
    idx_mklass = col_map["Materialklass"]
    idx_unit = col_map["Enhet"]

    out_rows = []
    num_rows = 0
    num_ok = 0
    num_skipped = 0

    for row in rows_iter:
        num_rows += 1
        if row is None:
            continue

        art_val = row[idx_art] if idx_art < len(row) else None
        gnp_val = row[idx_gnp] if idx_gnp < len(row) else None

        if art_val is None or gnp_val is None:
            num_skipped += 1
            continue

        # Artikelnummer som sträng med bibehållna ledande nollor
        artikelnummer = str(art_val).strip()

        # GNP → gn_pris (float)
        try:
            gn_pris = float(gnp_val)
        except (TypeError, ValueError):
            num_skipped += 1
            continue

        benamning = row[idx_ben] if idx_ben < len(row) else None
        materialklass = row[idx_mklass] if idx_mklass < len(row) else None
        enhet = row[idx_unit] if idx_unit < len(row) else None

        out_rows.append(
            {
                "artikelnummer": artikelnummer,
                "gn_pris": gn_pris,
                "benamning": str(benamning).strip() if benamning is not None else "",
                "materialklass": str(materialklass).strip() if materialklass is not None else "",
                "enhet": str(enhet).strip() if enhet is not None else "",
                "källa": "ahlsell_el_excel",
            }
        )
        num_ok += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(out_rows, f, ensure_ascii=False, indent=2)

    print(f"Antal datarader i Excel (exkl. header): {num_rows}")
    print(f"Antal rader exporterade: {num_ok}")
    print(f"Antal rader hoppade över (saknar artnr/pris eller ogiltigt pris): {num_skipped}")
    print(f"Skrev JSON till: {output_path}")


if __name__ == "__main__":
    main()
